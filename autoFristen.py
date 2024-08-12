
### BIBLIOTHEKEN
######################################################################################################################################################
from PySide6.QtWidgets import QApplication, QMainWindow, QMessageBox, QFileDialog
from frm_main import Ui_frm_main
from datetime import datetime, timedelta
from PySide6.QtCore import QLocale, Qt, QAbstractTableModel
from openpyxl.styles import Alignment
import pyodbc
import json
import locale
import pandas as pd
import os
import csv


### HAUPTKLASSE
######################################################################################################################################################
class Frm_main(QMainWindow, Ui_frm_main):


### HAUPTFUNKTION // NAVIGATION // BUTTONS
######################################################################################################################################################
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        locale.setlocale(locale.LC_TIME, 'de_DE.UTF-8')
        QLocale.setDefault(QLocale(QLocale.German))
        global strDatum
        global DatumMontag
        DatumMontag = (datetime.today().date() - timedelta(days=datetime.today().date().weekday()))

        self.conn = self.SQLconnection()
        self.myCursor = self.conn.cursor()
        
        ### MENÜ    ############################################################
        self.btn_home.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.pg_home))

        self.btn_create.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.pg_create))
        self.btn_create.clicked.connect(self.load_values)

        self.btn_createMA.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.pg_createMA))
        self.btn_createMA.clicked.connect(self.load_values)

        self.btn_createHS.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.pg_createHS))
        self.btn_createHS.clicked.connect(self.load_values)
        self.btn_createHS.clicked.connect(self.UpdateHSList)

        self.btn_doku.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.pg_doku))
        self.btn_doku.clicked.connect(self.load_csv)

        self.btn_mail.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.pg_mail))
        self.btn_mail.clicked.connect(self.load_Maildefault)

        self.btn_settings.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.pg_settings))
        self.btn_settings.clicked.connect(self.load_default)

        ### CREATE LISTS - Button    ###########################################
        self.btn_generate.clicked.connect(self.startCreationGENERAL)

        ### CREATE LISTS MITARBEITER - Button    ###########################################
        self.rb_InnoMgr.clicked.connect(self.UpdateMAList)
        self.rb_weitereMA.clicked.connect(self.UpdateMAList)
        self.rb_InnoMgr.clicked.connect(self.enableMAfunctions)
        self.rb_weitereMA.clicked.connect(self.enableMAfunctions)
        self.btn_generateMA.clicked.connect(self.startCreationMitarbeiter)

        ### CREATE LISTS HOCHSCHULEN - Button    ###########################################
        self.btn_generateHS.clicked.connect(self.startCreationHochschule)
        self.lb_DatumHS.dateChanged.connect(self.UpdateHSList)

        ### MAIL CONTENT - Button    ###########################################
        self.btn_save.clicked.connect(self.save_changesMail)

        ### EINSTELLUNGEN - Button   ###########################################
        self.btn_svchange.clicked.connect(self.save_changes)
        self.btn_pathma.clicked.connect(lambda: self.select_folder(self.df_pathIM))
        self.btn_pathdf.clicked.connect(lambda: self.select_folder(self.df_pathDF))
        self.btn_pathhs.clicked.connect(lambda: self.select_folder(self.df_pathHS))

### FUNKTIONEN FRISTENLISTEN
######################################################################################################################################################
    def startCreationGENERAL(self):
        self.ListType = 'general'
        self.create_List(self.cb_InnoMgr.isChecked(), self.cb_BackOffice.isChecked(), self.lb_Datum.text(), self.cb_Mail.isChecked())
    
    def startCreationMitarbeiter(self):
        self.ListType = 'Mitarbeiter'
        self.create_List(self.FLInnoMgr, self.FLweitereMA, self.lb_DatumMA.text(), self.cb_MailMA.isChecked())

    def startCreationHochschule(self):
        
        with open(self.get_default(), 'r') as f:
            self.constants = json.load(f)

        arrResponsible = self.getarrResponsibleHOCHSCHULE()

        self.start_creationHS(arrResponsible)

        self.systemnachricht("Fristenlisten wurden erstellt")

    def create_List(self, FLInnoMgr, FLweitereMA, lbDate, wMail):
        self.pb_Inno.setValue(0)
        self.pb_BackOffice.setValue(0)
        arrMail=[]
        with open(self.get_default(), 'r') as f:
            self.constants = json.load(f)

        if FLInnoMgr:
            if self.ListType == 'general':
                arrResponsible = self.getarrResponsibleGENERAL('SPALTENNAME AUS SQL Source1', lbDate, 'InnoMgr')                        #### 'SPALTENNAME AUS SQL' muss mit dem entsprechenden Namen der Spalte aus dem SQL-Query ausgetauscht werden. ###########################
            if self.ListType == 'Mitarbeiter':
                arrResponsible = self.getarrResponsibleMITARBEITER()
            
            self.start_creation(lbDate, self.constants.get('df_str_folderMA'), self.constants.get('df_str_folderdf'), 
                                self.pb_Inno, arrMail, 'InnoMgr', arrResponsible)
        if FLweitereMA:
            if self.ListType == 'general':
                arrResponsible = self.getarrResponsibleGENERAL('SPALTENNAME AUS SQL Source2', lbDate, 'SCR')                            #### 'SPALTENNAME AUS SQL' muss mit dem entsprechenden Namen der Spalte aus dem SQL-Query ausgetauscht werden. ###########################
            if self.ListType == 'Mitarbeiter':
                arrResponsible = self.getarrResponsibleMITARBEITER()

            self.start_creation(lbDate, self.constants.get('df_str_folderMA'), self.constants.get('df_str_folderdf'), 
                                self.pb_BackOffice, arrMail, 'SCR', arrResponsible)

        if len(arrMail) > 0 and wMail:
            self.send_Mail(arrMail, self.constants.get('df_CC'), self.lb_Datum)
            self.systemnachricht("Die Fristenlisten wurden erstellt und die Mail vorbereitet!")
        elif len(arrMail) > 0:
            self.systemnachricht("Die Fristenlisten wurden erstellt!")
        else:
            self.systemnachricht("Es gibt keine Fristenlisten für den angegebenen Zeitraum.")

    def start_creationHS(self, arrResponsible):
    
        date_obj = datetime.strptime(self.lb_DatumHS.text(), "%d.%m.%Y")
        formatted_date = date_obj.strftime("%Y-%m-%d")

        for HS in arrResponsible:
            DateNow = datetime.now().strftime("%Y.%m.%d - %H:%M")
            DateFile = str(datetime.now().strftime('%Y-%m-%d_%H%M%S'))

            Folder = self.constants.get('df_str_folderHS') + "\\" + DateFile + " - Fristenliste - " + HS + ".xlsx"

            df = pd.read_sql_query(self.getHSquery(HS, formatted_date), self.conn)

            df.to_excel(Folder, index=False)

            with pd.ExcelWriter(Folder, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Tabelle1')
                workbook = writer.book
                sheet = workbook['Tabelle1']
                
                sheet.column_dimensions['A'].width = 13     
                sheet.column_dimensions['B'].width = 13     
                sheet.column_dimensions['C'].width = 30    
                sheet.column_dimensions['D'].width = 30     
                sheet.column_dimensions['E'].width = 12     
                sheet.column_dimensions['F'].width = 12     
                sheet.column_dimensions['G'].width = 12     
                sheet.column_dimensions['H'].width = 30    
                sheet.column_dimensions['I'].width = 30     
                sheet.column_dimensions['J'].width = 30     
                sheet.column_dimensions['K'].width = 40     

                # Textumbruch für das gesamte Sheet aktivieren
                for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')

            self.append_csv(DateNow, HS, 'HS', Folder)

    def start_creation(self, Fdate, fMA, fdf, progress, EMail, ListT, arrResponsible):

        if not os.path.exists(fMA):
            self.systemnachricht("Der Dateipfad zu den Mitarbeiterordnern existiert nicht. Daher wird der Prozess abgebrochen.")
            return
        if not os.path.exists(fdf):
            self.systemnachricht("Der Dateipfad zum Default Ordner existiert nicht. Daher wird der Prozess abgebrochen.")
            return
     

        #Fdate = datetime.strptime(Fdate,"%d.%m.%Y")
        date_obj = datetime.strptime(Fdate, "%d.%m.%Y")
        formatted_date = date_obj.strftime("%Y-%m-%d")

        pStep = 100/len(arrResponsible)
        progress.setValue(0)
        i=0

        for strShort in arrResponsible:
            folder_MA = str(str(fMA) + "\\" + str(strShort))
            DateNow = datetime.now().strftime("%Y.%m.%d - %H:%M")
            DateFile = str(datetime.now().strftime('%Y-%m-%d_%H%M%S'))
            if not os.path.exists(folder_MA):
                folder_MA = fdf
            Folder = folder_MA
            Folder = Folder + "\\"
            Folder = Folder + DateFile
            Folder = Folder + " - Fristenliste - "
            Folder = Folder + ListT
            Folder = Folder + " - "
            Folder = Folder + strShort
            Folder = Folder + ".xlsx"

            if ListT == 'InnoMgr':
                df = pd.read_sql_query(self.getIMquery(strShort, formatted_date), self.conn)
            else:
                df = pd.read_sql_query(self.getBOquery(strShort, formatted_date), self.conn)

            df.to_excel(Folder, index=False)
        
            with pd.ExcelWriter(Folder, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Tabelle1')
                workbook = writer.book
                sheet = workbook['Tabelle1']
                
                sheet.column_dimensions['A'].width = 13    
                sheet.column_dimensions['B'].width = 13     
                sheet.column_dimensions['C'].width = 30     
                sheet.column_dimensions['D'].width = 30     
                sheet.column_dimensions['E'].width = 12     
                sheet.column_dimensions['F'].width = 12     
                sheet.column_dimensions['G'].width = 12     
                sheet.column_dimensions['H'].width = 30     
                sheet.column_dimensions['I'].width = 30     
                sheet.column_dimensions['J'].width = 30     
                sheet.column_dimensions['K'].width = 40     

                # Textumbruch für das gesamte Sheet aktivieren
                for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')

            self.append_csv(DateNow, strShort, ListT, Folder)
            EMail.append(strShort+"@provendis.info")
            i +=pStep
            progress.setValue(i)

    def send_Mail(self, arrMail, CCmail, FDate):
        from win32com.client import Dispatch                                    #Lazy import für Optimierung Initialisierung
        global strDatum
        global DatumMontag

        ### .txt öffnen für mail Subject                                        ######################################################################
        with open(self.get_defMailSubject(), 'r', encoding='utf-8') as fSubject:
            mailBetreff = fSubject.read()
        ### .txt öffnen für Mail content                                        ######################################################################
        with open(self.get_defMailContent(), 'r', encoding='utf-8') as fContenct:
            strBody = fContenct.read()
            
        strNext = DatumMontag + timedelta(days=14)
        strBearbeitung = strNext - timedelta(days=5)

        strNext = strNext.strftime("%d. %B '%y")
        strBearbeitung = strBearbeitung.strftime("%d. %B '%y")

        strBody = strBody.replace('[@Next]', strNext)
        strBody = strBody.replace('[@Bearbeitung]', strBearbeitung)
        mailBetreff = mailBetreff.replace('[@KW]', strDatum.strftime("KW %U/%Y"))

            
        arrMail=list(set(arrMail))
        strMail = "; ".join(arrMail)
        outlook = Dispatch('Outlook.Application')
        mail = outlook.CreateItem(0)
        mail.Subject = mailBetreff
        mail.To = strMail
        mail.Cc = CCmail
        mail.GetInspector  # Notwendig, um den Inspector abzurufen
        mail.HTMLBody = strBody + mail.HTMLBody  # Füge die Standardsignatur hinzu
    
        mail.display()

    def load_values(self):
        global strDatum
        with open(self.get_default(), 'r') as f:
            self.constants = json.load(f)
        dt_weeks = self.constants.get('df_int_Weeks')
        dt_Fristen = (datetime.today().date() - timedelta(days=datetime.today().date().weekday())) + timedelta(days=((7*dt_weeks)-1))
        self.lb_Datum.setDate(dt_Fristen)
        self.lb_DatumMA.setDate(dt_Fristen)
        self.lb_DatumHS.setDate(dt_Fristen)
        dt_Fristen = datetime.strftime(dt_Fristen, "%Y-%m-%d")
        strDatum = datetime.strptime(dt_Fristen, "%Y-%m-%d")

    def append_csv(self, datum, kuerzel, typ, quelle):

        # Öffnen der CSV-Datei im Anhangsmodus
        with open(self.get_doku(), mode='a', newline='', encoding='utf-8') as file:
            writer = csv.writer(file, delimiter=';')
            writer.writerow([datum, kuerzel, typ, quelle])
 
    def queryTOList(self, FDate, List):
        self.myCursor.execute(self.getResponsible(List, FDate))

        self.list_Mitarbeiter.clear()

        for row in sorted(self.myCursor.fetchall()):
            self.list_Mitarbeiter.addItem(row[0])

    def UpdateMAList(self):
        if self.rb_InnoMgr.isChecked():
            self.queryTOList(self.lb_DatumMA.text(), 'SPALTENNAME AUS SQL Source1')                            #### 'SPALTENNAME AUS SQL' muss mit dem entsprechenden Namen der Spalte aus dem SQL-Query ausgetauscht werden. ###########################
            self.FLInnoMgr = True
            self.FLweitereMA = False
        if self.rb_weitereMA.isChecked():
            self.queryTOList(self.lb_DatumMA.text(), 'SPALTENNAME AUS SQL Source2')                            #### 'SPALTENNAME AUS SQL' muss mit dem entsprechenden Namen der Spalte aus dem SQL-Query ausgetauscht werden. ###########################
            self.FLInnoMgr = False
            self.FLweitereMA = True

    def UpdateHSList(self):
        self.myCursor.execute(self.getResponsibleHS())

        self.list_Hochschulen.clear()

        for row in sorted(self.myCursor.fetchall()):
            self.list_Hochschulen.addItem(row[0])

    def enableMAfunctions(self):
        self.list_Mitarbeiter.setEnabled(True)
        self.cb_MailMA.setEnabled(True)
        self.btn_generateMA.setEnabled(True)

    def getarrResponsibleGENERAL(self, responsible, Fdate, ListT):

        df = pd.read_sql_query(self.getResponsible(responsible, Fdate),self.conn)

        if ListT == 'InnoMgr':
            arrResponsible = df['SPALTENNAME AUS SQL Source1'].tolist()                             #### 'SPALTENNAME AUS SQL' muss mit dem entsprechenden Namen der Spalte aus dem SQL-Query ausgetauscht werden OHNE TABELLENVERWEIS. ###########################
        else:
            arrResponsible = df['SPALTENNAME AUS SQL Source2'].tolist()                         #### 'SPALTENNAME AUS SQL' muss mit dem entsprechenden Namen der Spalte aus dem SQL-Query ausgetauscht werden OHNE TABELLENVERWEIS. ###########################

        return arrResponsible

    def getarrResponsibleMITARBEITER(self):
        arrResponsible = [item.text() for item in self.list_Mitarbeiter.selectedItems()]
        return arrResponsible
    
    def getarrResponsibleHOCHSCHULE(self):
        arrResponsible = [item.text() for item in self.list_Hochschulen.selectedItems()]
        return arrResponsible
       

### SQL DATENBANK FUNKTIONEN
######################################################################################################################################################
    def SQLconnection(self):
        conn = pyodbc.connect(
            'DRIVER={SQL Server};'
            'SERVER= *****;'                                           #### Servername eintragen. ###################################################################################################################################################
            'DATABASE= *****;'                                         #### Datenbankname eintragen. ################################################################################################################################################
            'Trusted_Connection=yes;'
            )
        return conn

    def getIMquery(self,strShort, FDate):
                                                    #### SQL QUERY für Datenbank ergänzen mit {strShort} als Platzhalter für den Responsible und {FDate} für Datum der zeitlichen Begrenzung . ####################################################################################################
        sql_query = f''' 
            SELECT

            FROM

            WHERE
                     = '{strShort}'
                     <= '{FDate}'

            ORDER BY

            '''
        return sql_query
    
    def getBOquery(self,strShort, FDate):
                                                    #### SQL QUERY für Datenbank ergänzen mit {strShort} als Platzhalter für den Responsible und {FDate} für Datum der zeitlichen Begrenzung . ##############################################################################
        sql_query = f'''
            SELECT

            FROM

            WHERE
                 = '{strShort}'
                 <= '{FDate}'

            ORDER BY
            
            '''
        return sql_query
 
    def getHSquery(self, HS, FDate):
                                                    #### SQL QUERY für Datenbank ergänzen mit {HS} als Platzhalter für den Stakeholder und {FDate} für Datum der zeitlichen Begrenzung . ##############################################################################
        sql_query = f'''
            SELECT

            FROM

            WHERE
                 = '{HS}'
                 <= '{FDate}'

            ORDER BY
            
            '''
        return sql_query        

    def getResponsible(self,List, FDate):
        sql_query = f'''
            SELECT DISTINCT
                {List}
            FROM
                Datenbank oder View Name
            WHERE
                     < '{FDate}'
                AND {List} <> 'NULL'
            '''
        return sql_query
   
    def getResponsibleHS(self):
        sql_query = f'''
            SELECT DISTINCT
                Spalte für Responsible
            FROM
                Tabellen oder View Name
            WHERE
                     < '{self.lb_DatumHS.text()}'
                AND Spalte für Responsible <> 'NULL'
            '''
        return sql_query


### FUNKTIONEN DOKUMENTATION
######################################################################################################################################################
    def load_csv(self):                                                                                                 # Läd die Inhalte der .csv in die Dokumentationstabelle
        ### .csv auslesen   ####################################################
        df = pd.read_csv(self.get_doku(), sep=';')

        ### Nur die ersten 4 Spalten auswählen  ################################
        if len(df.columns) >= 4:
            df = df.iloc[:, :4]

        ### Inhalt in Tabelle definieren und einfügen   ########################
        model = PandasModel(df)
        self.tb_doku.setModel(model)

        #### Spaltenbreiten definieren  ########################################
        self.tb_doku.setColumnWidth(0,140)                                      # Zeitstempel
        self.tb_doku.setColumnWidth(1,70)                                       # Mitarbeiterkürzel
        self.tb_doku.setColumnWidth(2,80)                                       # Team
        self.tb_doku.setColumnWidth(3,580)                                      # Speicherort


### FUNKTIONEN FÜR E-MAIL EINSTELLUNGEN
######################################################################################################################################################
    def load_Maildefault(self):                                                                                         # Läd die Default-Werte aus den .txt. Für Seitenaufruf

        ### .txt öffnen für mail Subject                                        ######################################################################
        with open(self.get_defMailSubject(), 'r', encoding='utf-8') as fSubject:
            self.txt_subject.setText(fSubject.read())

        ### .txt öffnen für Mail content                                        ######################################################################
        with open(self.get_defMailContent(), 'r', encoding='utf-8') as fContenct:
            self.txt_content.setPlainText(fContenct.read())

    def save_changesMail(self):                                                                                         # Änderungen in .txt speichern

        ### Speichern der geänderten Konstanten zurück in die JSON-Datei        ######################################################################
        with open(self.get_defMailSubject(), 'w', encoding='utf-8') as fSubject:
            fSubject.write(self.txt_subject.text())
        with open(self.get_defMailContent(), 'w', encoding='utf-8') as fContent:
            fContent.write(self.txt_content.toPlainText())
        
        ### Ausgabe Systemnachricht                                             ######################################################################
        self.systemnachricht("Default Werte gespeichert")


### FUNKTIONEN FÜR EINSTELLUNGEN
######################################################################################################################################################
    def load_default(self):                                                                                             # Läd die Default-Werte aus der .json. Für Seitenaufruf
        with open(self.get_default(), 'r') as f:
            self.constants = json.load(f)
        self.df_weeks.setValue(self.constants.get('df_int_Weeks'))
        self.df_pathIM.setText(self.constants.get('df_str_folderMA'))
        self.df_pathDF.setText(self.constants.get('df_str_folderdf'))
        self.df_pathHS.setText(self.constants.get('df_str_folderHS'))

        self.df_CC.setText(self.constants.get('df_CC'))

    def save_changes(self):                                                                                             # Änderungen in .json Datei speichern
        self.constants['df_int_Weeks'] = self.df_weeks.value()
        self.constants['df_str_folderMA'] = self.df_pathIM.text()
        self.constants['df_str_folderdf'] = self.df_pathDF.text()
        self.constants['df_str_folderHS'] = self.df_pathHS.text()

        self.constants['df_CC'] =self.df_CC.text()

        # Speichern der geänderten Konstanten zurück in die JSON-Datei
        with open(self.get_default(), 'w') as f:
            json.dump(self.constants, f)
        
        self.systemnachricht("Default Werte gespeichert")

    def select_folder(self,label):
        options = QFileDialog.Options()
        folder_path = QFileDialog.getExistingDirectory(self, "Ordner auswählen", options=options)
        if folder_path:
            label.setText(folder_path)

### SPEICHERORTE FÜR CONFIGURATIONEN ABHOLEN
######################################################################################################################################################
    def get_doku(self):                                                                                                 # Dokumentation der erstellten Listen (.csv)
        csv_filename = "config\\dokumentation.csv"
        return csv_filename  

    def get_defMailContent(self):                                                                                       # Content der E-Mail (.txt)                      
        txt_filename = "config\\mail.txt"
        return txt_filename

    def get_defMailSubject(self):                                                                                       # Subject der E-Mail (.txt)                      
        txt_filename = "config\\subject.txt"
        return txt_filename

    def get_default(self):                                                                                              # Default Einstellungen (.json)
        csv_filename = "config\\constants.json"
        return csv_filename


### ALLGEMEINE FUNKTIONEN
######################################################################################################################################################
    def systemnachricht(self,msg):                                                                                      # Erzeug Systemnachricht mit dem Inhalt msg
        msg_box = QMessageBox()
        msg_box.setWindowTitle("Systemnachricht")
        msg_box.setText(msg)
        msg_box.exec()
    
    def closeEvent(self, event):
        self.myCursor.close()
        self.conn.close()
        event.accept() 


### ERSTELLUNG TABELLENANSICHT DER DOKUMENTATION
######################################################################################################################################################
class PandasModel(QAbstractTableModel):
    def __init__(self, df):
        QAbstractTableModel.__init__(self)
        self._df = df

    def rowCount(self, parent=None):
        return len(self._df)

    def columnCount(self, parent=None):
        return len(self._df.columns)

    def data(self, index, role=Qt.DisplayRole):
        if index.isValid():
            if role == Qt.DisplayRole:
                return str(self._df.iloc[index.row(), index.column()])
        return None

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role == Qt.DisplayRole:
            if orientation == Qt.Horizontal:
                return self._df.columns[section]
            if orientation == Qt.Vertical:
                return str(self._df.index[section])
        return None


### AUSFÜHRUNG
######################################################################################################################################################    
app = QApplication()
frm_main = Frm_main()
frm_main.show()
app.exec()
